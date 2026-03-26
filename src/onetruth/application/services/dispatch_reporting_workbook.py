from __future__ import annotations

from collections.abc import Mapping, Sequence
from dataclasses import dataclass
import io
from typing import Any


WORKFLOW_ID = "dispatch_reporting.v1"
DATASET_KEY = "reporting.upd_draft.workbook"


@dataclass(frozen=True)
class WorkbookTableSpec:
    projection_key: str
    workbook_table_name: str
    columns: tuple[tuple[str, str], ...]
    editable: bool
    required: bool = True
    single_row: bool = False

    @property
    def headers(self) -> tuple[str, ...]:
        return tuple(header for header, _ in self.columns)

    @property
    def field_keys(self) -> tuple[str, ...]:
        return tuple(field_key for _, field_key in self.columns)


TABLE_SPECS: tuple[WorkbookTableSpec, ...] = (
    WorkbookTableSpec(
        projection_key="route_actuals",
        workbook_table_name="RouteActuals",
        columns=(
            ("RowID", "row_id"),
            ("ServiceDate", "service_date"),
            ("RouteID", "route_id"),
            ("DriverName", "driver_name"),
            ("PackagesDispatched", "packages_dispatched"),
            ("PackagesDelivered", "packages_delivered"),
            ("PlannedStart", "planned_start"),
            ("PlannedFinish", "planned_finish"),
            ("ActualStart", "actual_start"),
            ("ActualFinish", "actual_finish"),
            ("ActualMinutes", "actual_minutes"),
            ("Returns", "returns"),
            ("ReturnReasons", "return_reasons"),
            ("UpdCandidate", "upd_candidate"),
        ),
        editable=True,
    ),
    WorkbookTableSpec(
        projection_key="upd_candidates",
        workbook_table_name="UpdCandidates",
        columns=(
            ("RowID", "row_id"),
            ("ServiceDate", "service_date"),
            ("RouteID", "route_id"),
            ("DriverName", "driver_name"),
            ("ActualMinutes", "actual_minutes"),
            ("Selected", "selected"),
            ("Reason", "reason"),
            ("ManagerNote", "manager_note"),
        ),
        editable=True,
    ),
    WorkbookTableSpec(
        projection_key="manual_closeout",
        workbook_table_name="ManualCloseout",
        columns=(
            ("RowID", "row_id"),
            ("SickCalls", "sick_calls"),
            ("UnavailableDrivers", "unavailable_drivers"),
            ("WorkingDevices", "working_devices"),
            ("Rescues", "rescues"),
            ("Incidents", "incidents"),
            ("LastDriverClockout", "last_driver_clockout"),
            ("DispatcherComment", "dispatcher_comment"),
            ("ManagerNote", "manager_note"),
        ),
        editable=True,
        single_row=True,
    ),
    WorkbookTableSpec(
        projection_key="quality_warnings",
        workbook_table_name="QualityWarnings",
        columns=(
            ("RowID", "row_id"),
            ("WarningCode", "warning_code"),
            ("Severity", "severity"),
            ("Message", "message"),
            ("SourceSheet", "source_sheet"),
        ),
        editable=False,
    ),
    WorkbookTableSpec(
        projection_key="change_log_stage03_upd_draft",
        workbook_table_name="ChangeLogStage03_UpdDraft",
        columns=(
            ("RowID", "row_id"),
            ("ChangeType", "change_type"),
            ("ActorID", "actor_id"),
            ("ChangedAt", "changed_at"),
            ("Summary", "summary"),
        ),
        editable=False,
    ),
    WorkbookTableSpec(
        projection_key="lookups03",
        workbook_table_name="Lookups03",
        columns=(
            ("RowID", "row_id"),
            ("ListName", "list_name"),
            ("Value", "value"),
            ("Label", "label"),
        ),
        editable=False,
        required=False,
    ),
)

TABLE_SPECS_BY_KEY = {spec.projection_key: spec for spec in TABLE_SPECS}
EDITABLE_TABLE_KEYS = tuple(spec.projection_key for spec in TABLE_SPECS if spec.editable)
READ_ONLY_TABLE_KEYS = tuple(spec.projection_key for spec in TABLE_SPECS if not spec.editable)
SERVER_MANAGED_TABLE_KEY = "change_log_stage03_upd_draft"


def project_upd_draft_workbook(content: bytes) -> dict[str, Any]:
    workbook = _load_workbook_from_bytes(content)
    projection = {
        "workflow_id": WORKFLOW_ID,
        "dataset_key": DATASET_KEY,
        "editable_tables": list(EDITABLE_TABLE_KEYS),
        "read_only_tables": list(READ_ONLY_TABLE_KEYS),
    }
    for spec in TABLE_SPECS:
        projection[spec.projection_key] = _read_table_rows(workbook, spec)
    return projection


def materialize_upd_draft_workbook(
    base_content: bytes,
    edits: Mapping[str, Any],
    *,
    change_log_entry: Mapping[str, Any] | None = None,
) -> bytes:
    if not isinstance(edits, Mapping):
        raise ValueError("workbook edits must be a mapping")

    requested_keys = {str(key) for key in edits}
    unknown_keys = requested_keys - set(TABLE_SPECS_BY_KEY)
    if unknown_keys:
        raise ValueError(f"unknown workbook tables: {sorted(unknown_keys)}")

    read_only_keys = requested_keys - set(EDITABLE_TABLE_KEYS)
    if read_only_keys:
        raise ValueError(f"read-only workbook tables cannot be edited: {sorted(read_only_keys)}")

    workbook = _load_workbook_from_bytes(base_content)
    for table_key in EDITABLE_TABLE_KEYS:
        if table_key not in edits:
            continue
        spec = TABLE_SPECS_BY_KEY[table_key]
        normalized_rows = _normalize_client_rows(edits[table_key], spec=spec)
        _write_table_rows(workbook, spec, normalized_rows)

    if change_log_entry is not None:
        change_log_spec = TABLE_SPECS_BY_KEY[SERVER_MANAGED_TABLE_KEY]
        existing_rows = _read_table_rows(workbook, change_log_spec)
        appended_row = _normalize_client_rows([change_log_entry], spec=change_log_spec)
        _write_table_rows(workbook, change_log_spec, existing_rows + appended_row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _read_table_rows(workbook, spec: WorkbookTableSpec) -> list[dict[str, Any]]:
    worksheet = workbook[spec.workbook_table_name] if spec.workbook_table_name in workbook.sheetnames else None
    if worksheet is None:
        if spec.required:
            raise ValueError(f"required worksheet missing: {spec.workbook_table_name}")
        return []

    table = worksheet.tables.get(spec.workbook_table_name)
    if table is None:
        if spec.required:
            raise ValueError(f"required table missing: {spec.workbook_table_name}")
        return []

    min_col, min_row, max_col, max_row = _range_boundaries(table.ref)
    headers = [
        worksheet.cell(row=min_row, column=column_index).value
        for column_index in range(min_col, max_col + 1)
    ]
    expected_headers = list(spec.headers)
    if headers != expected_headers:
        raise ValueError(
            f"unexpected workbook headers for {spec.workbook_table_name}: {headers!r}"
        )

    rows: list[dict[str, Any]] = []
    for row_index in range(min_row + 1, max_row + 1):
        raw_values = [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(min_col, max_col + 1)
        ]
        row = {
            field_key: _normalize_cell_value(value)
            for (_, field_key), value in zip(spec.columns, raw_values, strict=True)
        }
        if _row_is_blank(row.values()):
            continue
        rows.append(row)

    if spec.single_row and len(rows) > 1:
        raise ValueError(f"{spec.workbook_table_name} must contain exactly one row")
    return rows


def _write_table_rows(workbook, spec: WorkbookTableSpec, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook[spec.workbook_table_name]
    table = worksheet.tables[spec.workbook_table_name]

    if spec.single_row and len(rows) != 1:
        raise ValueError(f"{spec.workbook_table_name} must contain exactly one row")

    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(list(spec.headers))
    if rows:
        for row in rows:
            worksheet.append([row[field_key] for field_key in spec.field_keys])
        last_row = len(rows) + 1
    else:
        worksheet.append([None for _ in spec.headers])
        last_row = 2

    table.ref = f"A1:{_get_column_letter(len(spec.headers))}{last_row}"


def _load_workbook_from_bytes(content: bytes):
    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(content))


def _range_boundaries(reference: str) -> tuple[int, int, int, int]:
    from openpyxl.utils import range_boundaries

    return range_boundaries(reference)


def _get_column_letter(column_index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(column_index)


def _normalize_client_rows(raw_rows: Any, *, spec: WorkbookTableSpec) -> list[dict[str, Any]]:
    if not isinstance(raw_rows, Sequence) or isinstance(raw_rows, (str, bytes, bytearray)):
        raise ValueError(f"{spec.projection_key} edits must be a list of row mappings")

    normalized_rows: list[dict[str, Any]] = []
    seen_row_ids: set[str] = set()
    for index, row in enumerate(raw_rows):
        if not isinstance(row, Mapping):
            raise ValueError(f"{spec.projection_key} row must be a mapping: index={index}")
        missing_fields = [field_key for field_key in spec.field_keys if field_key not in row]
        if missing_fields:
            raise ValueError(
                f"{spec.projection_key} row missing fields: index={index} fields={missing_fields}"
            )
        extra_fields = sorted(str(field_key) for field_key in row if field_key not in spec.field_keys)
        if extra_fields:
            raise ValueError(
                f"{spec.projection_key} row has unexpected fields: index={index} fields={extra_fields}"
            )

        normalized = {
            field_key: _normalize_cell_value(row[field_key])
            for field_key in spec.field_keys
        }
        row_id = str(normalized["row_id"]).strip()
        if not row_id:
            raise ValueError(f"{spec.projection_key} row_id is required: index={index}")
        if row_id in seen_row_ids:
            raise ValueError(
                f"{spec.projection_key} row_id must be unique: row_id={row_id}"
            )
        seen_row_ids.add(row_id)
        normalized["row_id"] = row_id
        normalized_rows.append(normalized)

    return normalized_rows


def _normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _row_is_blank(values: Sequence[Any]) -> bool:
    return all(str(value).strip() == "" for value in values)
