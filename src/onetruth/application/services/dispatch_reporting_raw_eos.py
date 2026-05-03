from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
import io
import re
from datetime import datetime, time, timedelta
from typing import Any

from onetruth.application.services.dispatch_reporting_workbook import (
    WorkbookRuntimeDependencyError,
)


RAW_EOS_WORKBOOK_FAMILY = "dispatch_reporting.raw_eos.workbook"

_PRIMARY_SHEET_NAME = "DSP RTS RETURNS"
_BREAK_TRACKER_SHEET_NAME = "Break Tracker"
_ROUTE_ADHERENCE_SHEET_NAME = "Route Adherence"
_AMZL_USE_SHEET_NAME = "AMZL USE"
_EMAIL_TEMPLATE_SHEET_NAME = "Email Template"

_ROUTE_ID_PATTERN = re.compile(r"^[A-Z]{1,5}\d{1,4}[A-Z]?$")
_SERVICE_DATE_PATTERN = re.compile(r"(?P<date>\d{4}-\d{2}-\d{2})")
_STATION_CODE_PATTERN = re.compile(r"\bDVC\s*(?P<digits>\d+)\b", re.IGNORECASE)
_BROKEN_FORMULA_PATTERN = re.compile(r"#REF!|#VALUE!|#NAME\?|#N/A", re.IGNORECASE)

_RETURN_REASON_COLUMNS: tuple[tuple[str, int], ...] = (
    ("UTA", 9),
    ("BC", 10),
    ("FDD", 11),
    ("REJ./ DAMAGED", 12),
    ("LOCKER FULL", 13),
    ("NSL", 14),
    ("UTL", 15),
    ("OODT", 16),
    ("MISSING", 18),
)


@dataclass(frozen=True)
class RawEosProjection:
    service_date: str
    station_code: str
    dsp_name: str
    route_rows: list[dict[str, Any]]
    quality_warning_rows: list[dict[str, Any]]
    break_tracker_rows: list[dict[str, Any]]
    break_tracker_sheet_present: bool
    route_adherence_cycles: list[dict[str, Any]]
    route_adherence_sheet_present: bool

    @property
    def formula_integrity_warning(self) -> bool:
        return bool(self.quality_warning_rows)


def project_raw_eos_workbook(
    content: bytes,
    *,
    source_metadata_json: Mapping[str, Any] | None = None,
    source_file_name: str | None = None,
    fallback_service_date: str | None = None,
    fallback_station_code: str | None = None,
    fallback_dsp_name: str | None = None,
) -> RawEosProjection:
    workbook = _load_workbook_from_bytes(content, data_only=False)
    workbook_values = _load_workbook_from_bytes(content, data_only=True)

    primary_sheet = _require_sheet(workbook, _PRIMARY_SHEET_NAME)
    primary_values_sheet = _require_sheet(workbook_values, _PRIMARY_SHEET_NAME)

    banner_text = _cell_text(primary_values_sheet.cell(1, 1).value)
    station_code = _resolve_station_code(
        banner_text=banner_text,
        metadata_json=source_metadata_json,
        fallback_station_code=fallback_station_code,
    )
    dsp_name = _resolve_dsp_name(
        banner_text=banner_text,
        metadata_json=source_metadata_json,
        fallback_dsp_name=fallback_dsp_name,
    )

    route_rows = _parse_route_rows(primary_values_sheet)
    if not route_rows:
        raise ValueError("EOS workbook did not contain any route-level actual rows")

    service_date = _resolve_service_date(
        metadata_json=source_metadata_json,
        source_file_name=source_file_name,
        fallback_service_date=fallback_service_date,
    )

    for row in route_rows:
        row["service_date"] = service_date

    quality_warning_rows = _collect_quality_warning_rows(workbook)

    break_tracker_sheet = _optional_sheet(workbook_values, _BREAK_TRACKER_SHEET_NAME)
    route_adherence_sheet = _optional_sheet(workbook_values, _ROUTE_ADHERENCE_SHEET_NAME)

    return RawEosProjection(
        service_date=service_date,
        station_code=station_code,
        dsp_name=dsp_name,
        route_rows=route_rows,
        quality_warning_rows=quality_warning_rows,
        break_tracker_rows=_parse_break_tracker_rows(break_tracker_sheet),
        break_tracker_sheet_present=break_tracker_sheet is not None,
        route_adherence_cycles=_parse_route_adherence_cycles(route_adherence_sheet),
        route_adherence_sheet_present=route_adherence_sheet is not None,
    )


def _parse_route_rows(sheet) -> list[dict[str, Any]]:
    route_rows: list[dict[str, Any]] = []
    for row_index in range(1, int(sheet.max_row) + 1):
        route_id = _cell_text(sheet.cell(row_index, 1).value)
        if not route_id or not _ROUTE_ID_PATTERN.match(route_id):
            continue
        driver_raw = _cell_text(sheet.cell(row_index, 2).value)
        if not driver_raw:
            continue
        packages_dispatched = _coerce_int(sheet.cell(row_index, 3).value)
        actual_dispatched = _coerce_int(sheet.cell(row_index, 4).value)
        packages_delivered = _coerce_int(sheet.cell(row_index, 5).value)
        returned_packages = _coerce_int(sheet.cell(row_index, 8).value)
        planned_start = _time_string(sheet.cell(row_index, 20).value)
        planned_finish = _time_string(sheet.cell(row_index, 21).value)
        actual_start = _time_string(sheet.cell(row_index, 23).value)
        actual_finish = _time_string(sheet.cell(row_index, 24).value)
        actual_minutes = _duration_minutes(
            sheet.cell(row_index, 25).value,
            start_value=sheet.cell(row_index, 23).value,
            finish_value=sheet.cell(row_index, 24).value,
        )
        route_rows.append(
            {
                "row_id": f"route-{route_id.lower()}",
                "service_date": "",
                "route_id": route_id,
                "driver_name": _normalize_driver_display_name(driver_raw),
                "driver_name_raw": driver_raw,
                "packages_dispatched": packages_dispatched,
                "actual_dispatched": actual_dispatched,
                "packages_delivered": packages_delivered,
                "planned_start": planned_start,
                "planned_finish": planned_finish,
                "actual_start": actual_start,
                "actual_finish": actual_finish,
                "actual_minutes": actual_minutes,
                "returns": returned_packages,
                "return_reasons": _return_reasons_text(sheet, row_index),
            }
        )
    return route_rows


def _collect_quality_warning_rows(workbook) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []

    for name, defined_name in workbook.defined_names.items():
        attr_text = str(getattr(defined_name, "attr_text", "") or "")
        if _BROKEN_FORMULA_PATTERN.search(attr_text):
            warnings.append(
                _quality_warning_row(
                    code="named_range_broken",
                    source_sheet="workbook.defined_names",
                    message=f"Named range {name} contains a broken reference.",
                )
            )

    for sheet_name in (_AMZL_USE_SHEET_NAME, _EMAIL_TEMPLATE_SHEET_NAME):
        sheet = _optional_sheet(workbook, sheet_name)
        if sheet is None:
            continue
        found_broken_formula = False
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                if _BROKEN_FORMULA_PATTERN.search(cell.value):
                    found_broken_formula = True
                    break
            if found_broken_formula:
                break
        if found_broken_formula:
            warnings.append(
                _quality_warning_row(
                    code="formula_integrity_warning",
                    source_sheet=sheet.title.strip(),
                    message=f"{sheet.title.strip()} contains broken summary formulas.",
                )
            )

    deduped: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in warnings:
        key = (str(row["warning_code"]), str(row["source_sheet"]))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def _quality_warning_row(*, code: str, source_sheet: str, message: str) -> dict[str, Any]:
    return {
        "row_id": f"warning-{len(source_sheet)}-{code.lower()}",
        "warning_code": code,
        "severity": "warning",
        "message": message,
        "source_sheet": source_sheet,
    }


def _parse_break_tracker_rows(sheet) -> list[dict[str, Any]]:
    if sheet is None:
        return []
    rows: list[dict[str, Any]] = []
    for row_index in range(4, int(sheet.max_row) + 1):
        route_id = _cell_text(sheet.cell(row_index, 2).value)
        driver_name = _cell_text(sheet.cell(row_index, 3).value)
        if not route_id and not driver_name:
            continue
        rows.append(
            {
                "row_id": f"break-{row_index:03d}",
                "cycle": _cell_text(sheet.cell(row_index, 1).value),
                "route_id": route_id,
                "driver_name": driver_name,
                "first_break_type": _cell_text(sheet.cell(row_index, 4).value),
                "first_break_stop": _cell_text(sheet.cell(row_index, 5).value),
                "first_break_start": _time_string(sheet.cell(row_index, 6).value),
                "first_break_end": _time_string(sheet.cell(row_index, 7).value),
                "first_break_duration": _duration_string(sheet.cell(row_index, 8).value),
                "second_break_type": _cell_text(sheet.cell(row_index, 9).value),
                "second_break_stop": _cell_text(sheet.cell(row_index, 10).value),
                "second_break_start": _time_string(sheet.cell(row_index, 11).value),
                "second_break_end": _time_string(sheet.cell(row_index, 12).value),
                "second_break_duration": _duration_string(sheet.cell(row_index, 13).value),
            }
        )
    return rows


def _parse_route_adherence_cycles(sheet) -> list[dict[str, Any]]:
    if sheet is None:
        return []
    cycles: list[dict[str, Any]] = []
    row_index = 1
    max_row = int(sheet.max_row)
    while row_index <= max_row:
        cycle_label = _cell_text(sheet.cell(row_index, 1).value)
        if not cycle_label.startswith("Cycle"):
            row_index += 1
            continue
        header_row_index = row_index + 1
        header_values = [
            _header_label(sheet.cell(header_row_index, column_index).value)
            for column_index in range(1, int(sheet.max_column) + 1)
        ]
        columns: list[dict[str, str]] = []
        for column_index, label in enumerate(header_values, start=1):
            if not label:
                continue
            columns.append(
                {
                    "key": _adherence_column_key(label, column_index),
                    "label": label,
                }
            )
        rows: list[dict[str, Any]] = []
        data_row_index = header_row_index + 1
        while data_row_index <= max_row:
            next_cycle = _cell_text(sheet.cell(data_row_index, 1).value)
            if next_cycle.startswith("Cycle"):
                break
            values = [
                sheet.cell(data_row_index, column_index).value
                for column_index in range(1, len(columns) + 1)
            ]
            if any(value not in (None, "") for value in values):
                row = {"row_id": f"route-adherence-{cycle_label.lower().replace(' ', '-')}-{data_row_index:03d}"}
                has_route_context = False
                for column_offset, column in enumerate(columns, start=1):
                    value = values[column_offset - 1]
                    normalized = _route_adherence_cell_value(value)
                    row[column["key"]] = normalized
                    if column["key"] in {"driver_name", "route_id", "auto_suggested"} and normalized:
                        has_route_context = True
                if has_route_context:
                    rows.append(row)
            data_row_index += 1
        cycles.append(
            {
                "cycle_label": cycle_label,
                "columns": columns,
                "rows": rows,
            }
        )
        row_index = data_row_index
    return cycles


def _resolve_service_date(
    *,
    metadata_json: Mapping[str, Any] | None,
    source_file_name: str | None,
    fallback_service_date: str | None,
) -> str:
    metadata_value = _lookup_text(metadata_json, "service_date")
    if metadata_value:
        return metadata_value
    file_name = source_file_name or ""
    match = _SERVICE_DATE_PATTERN.search(file_name)
    if match is not None:
        return str(match.group("date"))
    if fallback_service_date:
        return fallback_service_date
    raise ValueError("service date could not be resolved for EOS workbook")


def _resolve_station_code(
    *,
    banner_text: str,
    metadata_json: Mapping[str, Any] | None,
    fallback_station_code: str | None,
) -> str:
    match = _STATION_CODE_PATTERN.search(banner_text)
    if match is not None:
        return f"DVC{match.group('digits')}"
    metadata_value = _lookup_text(metadata_json, "station_code")
    if metadata_value:
        return metadata_value
    return fallback_station_code or "DVC4"


def _resolve_dsp_name(
    *,
    banner_text: str,
    metadata_json: Mapping[str, Any] | None,
    fallback_dsp_name: str | None,
) -> str:
    if banner_text:
        station_match = _STATION_CODE_PATTERN.search(banner_text)
        if station_match is not None:
            prefix = banner_text[: station_match.start()].strip()
            if prefix:
                return prefix.replace("  ", " ")
    metadata_value = _lookup_text(metadata_json, "dsp_name")
    if metadata_value:
        return metadata_value
    return fallback_dsp_name or "QDCI"


def _return_reasons_text(sheet, row_index: int) -> str:
    reasons: list[str] = []
    for reason_label, column_index in _RETURN_REASON_COLUMNS:
        count = _coerce_int(sheet.cell(row_index, column_index).value)
        if count <= 0:
            continue
        reasons.append(f"{reason_label}:{count}")
    return ", ".join(reasons)


def _normalize_driver_display_name(value: str) -> str:
    head = value.split(" / ", 1)[0].strip()
    if "(" in head:
        head = head.split("(", 1)[0].strip()
    return head


def _header_label(value: Any) -> str:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return _cell_text(value)


def _adherence_column_key(label: str, column_index: int) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_")
    if normalized == "da_name":
        return "driver_name"
    if normalized == "auto_suggested":
        return "auto_suggested"
    if normalized == "route":
        return "route_id"
    if not normalized:
        return f"column_{column_index}"
    return normalized


def _route_adherence_cell_value(value: Any) -> Any:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    return _cell_text(value)


def _duration_minutes(value: Any, *, start_value: Any, finish_value: Any) -> int:
    direct = _coerce_duration_to_minutes(value)
    if direct is not None:
        return direct
    return _duration_between(start_value, finish_value)


def _duration_string(value: Any) -> str:
    minutes = _coerce_duration_to_minutes(value)
    if minutes is None:
        return ""
    hours, remainder = divmod(minutes, 60)
    return f"{hours:02d}:{remainder:02d}"


def _coerce_duration_to_minutes(value: Any) -> int | None:
    if isinstance(value, timedelta):
        return int(round(value.total_seconds() / 60))
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + (1 if value.second >= 30 else 0)
    if isinstance(value, time):
        return value.hour * 60 + value.minute + (1 if value.second >= 30 else 0)
    if isinstance(value, (int, float)):
        return int(round(float(value) * 24 * 60))
    return None


def _duration_between(start_value: Any, finish_value: Any) -> int:
    start = _minutes_of_day(start_value)
    finish = _minutes_of_day(finish_value)
    if start is None or finish is None:
        return 0
    if finish < start:
        finish += 24 * 60
    return finish - start


def _minutes_of_day(value: Any) -> int | None:
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    return None


def _time_string(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return _cell_text(value)


def _coerce_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = _cell_text(value)
    if not text:
        return 0
    return int(round(float(text)))


def _lookup_text(metadata_json: Mapping[str, Any] | None, key: str) -> str:
    if not isinstance(metadata_json, Mapping):
        return ""
    return _cell_text(metadata_json.get(key))


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _require_sheet(workbook, normalized_name: str):
    sheet = _optional_sheet(workbook, normalized_name)
    if sheet is None:
        raise ValueError(f"required worksheet missing: {normalized_name}")
    return sheet


def _optional_sheet(workbook, normalized_name: str):
    wanted = normalized_name.strip().lower()
    for sheet in workbook.worksheets:
        if str(sheet.title).strip().lower() == wanted:
            return sheet
    return None


def _load_workbook_from_bytes(content: bytes, *, data_only: bool):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        if _is_missing_openpyxl(exc):
            raise WorkbookRuntimeDependencyError("openpyxl") from exc
        raise

    return load_workbook(io.BytesIO(content), data_only=data_only)


def _is_missing_openpyxl(exc: ModuleNotFoundError) -> bool:
    return exc.name == "openpyxl"
